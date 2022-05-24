import openpyxl

wb = openpyxl.Workbook()

# 上書き保存（読み込んだのと同じ名前を指定）
wb.save("Sample.xlsx")

# 名前が「Sheet1」のシート
ws = wb["Sheet"]

# 先頭のシート（注意：インデックス番号は0から始まる）
ws = wb.worksheets[0]

# インデックス番号の確認
index = wb.index(ws)
print( "index=", index)


# シート名のリスト
print( "sheetnames=",wb.sheetnames)
#['Sheet1', 'Sheet2', 'Sheet3']

# シート名の確認
print( "sheettitle=", ws.title)
#'Sheet1'

# シート名の変更
ws.title = "Sheet1"
print( "sheetnames=", wb.sheetnames)
#['SheetOne', 'Sheet2', 'Sheet3']


# 「Sheet2」のコピーをシートの最後に追加
ws2_copy = wb.copy_worksheet(wb["Sheet1"])
print( "add copy=", wb.sheetnames)

# 「Sheet4」を末尾に追加
ws4 = wb.create_sheet(title="Sheet4")
print( "add sheet4=", wb.sheetnames)
#['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']

# 「New Sheet」を左から３つ目の位置に追加
ws_new = wb.create_sheet(title="New Sheet", index=2)
print( "insert", wb.sheetnames)
#['Sheet1', 'Sheet2', 'New Sheet', 'Sheet3', 'Sheet4']


# 「Sheet2」のコピーを削除
wb.remove(ws2_copy)
print( "delete copy=", wb.sheetnames)


# 末尾のシートを削除
wb.remove(wb.worksheets[-1])
print( "delete tail=",wb.sheetnames)

